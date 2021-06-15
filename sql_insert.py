from models.models import *
from openpyxl import load_workbook

# wb = load_workbook('dhh.xlsx')
# sheet = wb[f'SkuChangeRequests']
# mass = {}
# m_row = sheet.max_row
engine = database()
# for j in range(1, 10):
# 	mass_test = []
# 	for i in range(19, 23):
# 	    cell_obj = sheet.cell(row = j, column = i)
# 	    if cell_obj.value not in mass_test:
# 	    	mass_test.append(cell_obj.value)
# 	mass[j] = mass_test

# bd_Session = sessionmaker(bind=engine)
# bd_session = bd_Session()
# count = 0
with open('категории.txt', encoding='utf-8', newline='') as f:
	my_lines = f.readlines()
	for key in my_lines:
		print(key, end='')
		bd_Session = sessionmaker(bind=engine)
		bd_session = bd_Session()
		category = Category(name=key)
		bd_session.add(category)
		bd_session.commit()
	bd_session.close()
# for i in range(1, 10):
# 	self.dictionary[i - 1] = []
# 	for j in range(1, 10):
# 		self.dictionary[i - 1].append(sheet.cell(row=i, column=j).value)